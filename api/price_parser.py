"""단가표 엑셀 파서 — supplier_presets.rules(JSONB)를 소비하는 순수 함수 (DB 무관).

T6(ERD §8) 첫 단계 실현: 실파일 .xlsx → supplier_price_rows 계약 행 목록.
출력 행 = {model_name, danawa_code, prices(JSONB dict), cost_price, supply_state, memo}
— admin_price_import._file_diff가 이 스키마에만 의존하므로 이후 diff/apply/undo 무변경.

rules 소비 키(seed_0008에서 확장):
  sheets.exclude   : 제외 시트명(빈/정책/이질 레이아웃 — MSI 저장장치·파워-쿨러는 v2 이관)
  model_col        : 모델명 헤더 라벨 = 헤더 앵커(상단 12행에서 탐지 — GBT POWER 2+3행·
                     MSI AMD 2+3행 등 시트별 오프셋 편차를 흡수, 하드코딩 회피)
  model_col_fallback: 모델명 결측 시 대체 열(MSI: MKT NAME)
  danawa_code_col  : 다나와 상품코드 열(GBT만)
  state_col/memo_col: 상태·비고 열
  state_map        : 정확 일치 정규화(O→가능 등). 미일치 시 키워드 스캔(품절→품절·문의→문의,
                     기본 가능 — MSI '단종예정'은 가능, 원문은 memo 보존). memo에 '문의' 포함 시
                     최종 override(사람 확인 필요 신호).
  price_cols       : cost_price = 이 열들의 유효값(>0) 최저(cost_rule=min_price)
  carry_forward    : 세로 병합 그룹 열(칩셋) — forward-fill(병합 셀은 좌상단만 값)

정규화: 헤더 키·모델명·코드에서 \\xa0(nbsp)→공백, strip, 다중 공백 축약. prices JSONB 키는
헤더에서 공백·개행 전부 제거("판매가\\n(윈윈, 셀프로)"→"판매가(윈윈,셀프로)" — 기존 rules
price_cols 표기와 정합). 가격 0·빈 셀은 무효(0원은 유효가가 아님).
이관: 원본 파일 보관, 해시 중복 검출, MSI 저장장치·파워-쿨러 레이아웃(v2), model_key 정규화.
"""
import io
import re

import openpyxl

HEADER_SCAN_ROWS = 12   # 헤더 앵커 탐지 범위
EMPTY_ROW_STOP = 30     # 연속 무효 행이 이만큼이면 표 종료로 판정(하단 거대 빈 영역 회피)


def _norm(v) -> str:
    """셀 텍스트 정규화 — nbsp 제거·strip·다중 공백 축약."""
    if v is None:
        return ""
    s = str(v).replace(" ", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()


def _hkey(v) -> str:
    """헤더 키 정규화 — 공백·개행 전부 제거(rules price_cols 표기와 정합)."""
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v).replace(" ", ""))


def _int_or_none(v):
    if v is None or isinstance(v, str):
        return None
    try:
        n = int(round(float(v)))
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _code_str(v) -> str | None:
    s = _norm(v)
    if not s:
        return None
    try:
        return str(int(float(s)))  # 69699425.0 → '69699425'
    except ValueError:
        return s[:20]


def _resolve_state(state_text: str, memo_text: str, state_map: dict) -> str:
    st = state_map.get(state_text)
    if st is None:
        st = "품절" if "품절" in state_text else ("문의" if "문의" in state_text else "가능")
    if "문의" in memo_text:  # 비고의 '문의'는 최종 override — 사람 확인 신호
        st = "문의"
    return st


def _parse_sheet(ws, rules: dict):
    """단일 시트 → (rows, skipped_rows) 또는 (None, 사유)."""
    model_col = rules["model_col"]
    fallback = rules.get("model_col_fallback")
    # 1) 헤더 앵커 탐지: model_col 라벨이 있는 행
    anchor_row = anchor_col = None
    for row in ws.iter_rows(min_row=1, max_row=min(HEADER_SCAN_ROWS, ws.max_row)):
        for cell in row:
            if _norm(cell.value) == model_col:
                anchor_row, anchor_col = cell.row, cell.column
                break
        if anchor_row:
            break
    if anchor_row is None:
        return None, "헤더 인식 실패"
    # 2) 열 맵: 주 헤더(앵커 행) + 부 헤더(다음 행) 합성 키
    headers, used = {}, set()
    for cell in ws[anchor_row]:
        main = _hkey(cell.value)
        sub = _hkey(ws.cell(row=anchor_row + 1, column=cell.column).value)
        key = main + sub if (main and sub and sub != main) else (main or sub)
        if key:
            # 병합 주헤더의 비앵커 열은 부헤더만 남아 키가 겹칠 수 있음(GBT 가상/미니샵 '카드가') — 접미사 dedupe
            while key in used:
                key += "_"
            used.add(key)
            headers[cell.column] = key
    def col_of(label):
        want = _hkey(label)
        return next((c for c, k in headers.items() if k == want), None)
    c_model, c_fb = col_of(model_col), (col_of(fallback) if fallback else None)
    c_state, c_memo = col_of(rules.get("state_col", "")), col_of(rules.get("memo_col", ""))
    c_danawa = col_of(rules.get("danawa_code_col", ""))
    c_carry = next((col_of(c) for c in rules.get("carry_forward", []) if col_of(c)), None)
    price_keys = [_hkey(c) for c in rules.get("price_cols", [])]
    meta_cols = {c for c in (c_model, c_fb, c_state, c_memo, c_danawa, c_carry) if c}

    rows, skipped, empty_streak, carry = [], 0, 0, None
    for row in ws.iter_rows(min_row=anchor_row + 2, max_row=ws.max_row):
        cells = {c.column: c.value for c in row}
        if c_carry and _norm(cells.get(c_carry)):
            carry = _norm(cells.get(c_carry))  # 병합 그룹 forward-fill
        model = _norm(cells.get(c_model)) or (_norm(cells.get(c_fb)) if c_fb else "")
        prices = {}
        for col, key in headers.items():
            if col in meta_cols:
                continue
            n = _int_or_none(cells.get(col))
            if n is not None:
                prices[key] = n
        valid = [prices[k] for k in price_keys if k in prices]
        if not model or model in ("\\", "0") or not valid:
            skipped += 1 if (model or prices) else 0  # 완전 빈 행은 스킵 카운트 제외
            empty_streak += 1
            if empty_streak >= EMPTY_ROW_STOP:
                break
            continue
        empty_streak = 0
        state_text = _norm(cells.get(c_state)) if c_state else ""
        memo_text = _norm(cells.get(c_memo)) if c_memo else ""
        rows.append({
            "model_name": model[:300],
            "danawa_code": _code_str(cells.get(c_danawa)) if c_danawa else None,
            "prices": prices,
            "cost_price": min(valid),  # cost_rule = min_price
            "supply_state": _resolve_state(state_text, memo_text, rules.get("state_map", {})),
            "memo": (memo_text[:200] or None),
            "chipset": carry,  # 참고용(스냅샷 컬럼엔 미저장 — memo 아님)
        })
    return rows, skipped


def parse_price_file(data: bytes, rules: dict) -> dict:
    """xlsx 바이트 → {rows, sheets:[{name,rows}], skipped_sheets:[{name,reason}], skipped_rows}."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    exclude = set((rules.get("sheets") or {}).get("exclude", []))
    all_rows, sheets, skipped_sheets, skipped_rows = [], [], [], 0
    for ws in wb.worksheets:
        name = ws.title.strip()
        if ws.title in exclude or name in exclude:
            skipped_sheets.append({"name": ws.title, "reason": "프리셋 제외"})
            continue
        if ws.sheet_state != "visible":
            skipped_sheets.append({"name": ws.title, "reason": "숨김 시트"})
            continue
        result, info = _parse_sheet(ws, rules)
        if result is None:
            skipped_sheets.append({"name": ws.title, "reason": info})
            continue
        sheets.append({"name": ws.title, "rows": len(result)})
        all_rows.extend(result)
        skipped_rows += info
    return {"rows": all_rows, "sheets": sheets,
            "skipped_sheets": skipped_sheets, "skipped_rows": skipped_rows}
